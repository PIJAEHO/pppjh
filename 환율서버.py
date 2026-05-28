from flask import Flask, jsonify, render_template_string, request, send_file
import requests
import urllib3
from datetime import datetime, timedelta
import io
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup
import time

urllib3.disable_warnings()

app = Flask(__name__)

API_KEY = "5DX7W4ohTJqIg9EZRQXnSfEpPqzi9yEu"
TARGET_CURRENCIES = ["USD", "EUR"]

HTML = """
<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>환율 · 금속 · 유가</title>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: 'Segoe UI', sans-serif; background: #f0f2f5; min-height: 100vh; display: flex; flex-direction: column; align-items: center; padding: 40px 16px; }
    h1 { font-size: 1.6rem; font-weight: 700; color: #1a1a2e; margin-bottom: 24px; }
    .panel { background: white; border-radius: 16px; padding: 24px 28px; width: 100%; max-width: 1100px; box-shadow: 0 4px 20px rgba(0,0,0,0.08); margin-bottom: 24px; }
    .row { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
    input[type=date] { padding: 8px 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 0.95rem; }
    .btn { padding: 8px 20px; border: none; border-radius: 8px; font-size: 0.95rem; cursor: pointer; }
    .btn-primary { background: #2563eb; color: white; }
    .btn-primary:hover { background: #1d4ed8; }
    .btn-green { background: #16a34a; color: white; }
    .btn-green:hover { background: #15803d; }
    .btn:disabled { background: #ccc; cursor: not-allowed; }
    #status { font-size: 0.85rem; color: #888; margin-top: 12px; }
    #error { color: #dc2626; margin-top: 12px; font-size: 0.9rem; }
    #tableWrap { margin-top: 20px; overflow-x: auto; }
    table { width: 100%; border-collapse: collapse; font-size: 0.88rem; }
    .th-top { color: white; font-weight: 700; text-align: center; padding: 10px; }
    .th-top.fx  { background: #1a3a5c; }
    .th-top.lme { background: #3b1a5c; }
    .th-top.oil { background: #1a4a2e; }
    .th-sub { color: white; font-weight: 600; padding: 8px 10px; text-align: center; border: 1px solid rgba(255,255,255,0.2); }
    .th-sub.fx  { background: #2d5f8a; }
    .th-sub.lme { background: #6b2d8a; }
    .th-sub.oil { background: #2d7a50; }
    .th-date { background: #2d4a6b; color: white; font-weight: 600; padding: 8px 12px; text-align: center; }
    td { padding: 7px 12px; border: 1px solid #e5e7eb; text-align: right; color: #222; }
    td:first-child { text-align: center; background: #f8fafc; color: #555; font-weight: 500; }
    tr:hover td { background: #f0f6f2; }
    tr:hover td:first-child { background: #e8f0eb; }
  </style>
</head>
<body>
  <h1>환율 · 금속 · 유가 시세</h1>
  <div class="panel">
    <div class="row">
      <input type="date" id="startDate">
      <span style="color:#888;">~</span>
      <input type="date" id="endDate">
      <button class="btn btn-primary" onclick="loadRange()">조회</button>
      <button class="btn btn-green" id="downloadBtn" onclick="downloadExcel()" disabled>엑셀 다운로드</button>
    </div>
    <div id="status"></div>
    <div id="error"></div>
    <div id="tableWrap"></div>
  </div>

  <script>
    const today = new Date().toISOString().slice(0, 10);
    const weekAgo = new Date(Date.now() - 7 * 86400000).toISOString().slice(0, 10);
    document.getElementById('startDate').value = weekAgo;
    document.getElementById('endDate').value = today;

    async function loadRange() {
      const start = document.getElementById('startDate').value.replace(/-/g, '');
      const end = document.getElementById('endDate').value.replace(/-/g, '');
      if (!start || !end || start > end) {
        document.getElementById('error').textContent = '날짜 범위를 올바르게 입력해주세요.';
        return;
      }
      document.getElementById('error').textContent = '';
      document.getElementById('status').textContent = '조회 중... (환율 + LME + 유가 데이터 수집 중)';
      document.getElementById('tableWrap').innerHTML = '';
      document.getElementById('downloadBtn').disabled = true;

      try {
        const res = await fetch('/api/range?start=' + start + '&end=' + end);
        const data = await res.json();
        if (data.error) { document.getElementById('error').textContent = '오류: ' + data.error; return; }

        document.getElementById('status').textContent = '총 ' + data.rows.length + '일 조회됨';
        document.getElementById('downloadBtn').disabled = false;

        let html = '<table><thead>'
          + '<tr>'
          + '<th class="th-date" rowspan="2">날짜</th>'
          + '<th class="th-top fx" colspan="2">환율 (KRW)</th>'
          + '<th class="th-top lme" colspan="2">LME 현물 (US$/톤)</th>'
          + '<th class="th-top oil" colspan="3">국제유가 ($/Bbl)</th>'
          + '</tr>'
          + '<tr>'
          + '<th class="th-sub fx">USD/KRW</th>'
          + '<th class="th-sub fx">EUR/KRW</th>'
          + '<th class="th-sub lme">니켈 (Ni)</th>'
          + '<th class="th-sub lme">구리 (Cu)</th>'
          + '<th class="th-sub oil">두바이유</th>'
          + '<th class="th-sub oil">브렌트</th>'
          + '<th class="th-sub oil">WTI</th>'
          + '</tr></thead><tbody>';

        data.rows.forEach(r => {
          const d = r.date.slice(0,4) + '-' + r.date.slice(4,6) + '-' + r.date.slice(6,8);
          html += '<tr>'
            + '<td>' + d + '</td>'
            + '<td>' + (r.USD   || '-') + '</td>'
            + '<td>' + (r.EUR   || '-') + '</td>'
            + '<td>' + (r.Ni    || '-') + '</td>'
            + '<td>' + (r.Cu    || '-') + '</td>'
            + '<td>' + (r.Dubai || '-') + '</td>'
            + '<td>' + (r.Brent || '-') + '</td>'
            + '<td>' + (r.WTI   || '-') + '</td>'
            + '</tr>';
        });
        html += '</tbody></table>';
        document.getElementById('tableWrap').innerHTML = html;
      } catch(e) {
        document.getElementById('error').textContent = '서버 오류: ' + e.message;
      }
    }

    function downloadExcel() {
      const start = document.getElementById('startDate').value.replace(/-/g, '');
      const end = document.getElementById('endDate').value.replace(/-/g, '');
      window.location.href = '/api/download?start=' + start + '&end=' + end;
    }
  </script>
</body>
</html>
"""


def fetch_one_day_fx(date: str) -> list:
    url = "https://oapi.koreaexim.go.kr/site/program/financial/exchangeJSON"
    params = {"authkey": API_KEY, "searchdate": date, "data": "AP01"}
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    for attempt in range(3):
        try:
            resp = requests.get(url, params=params, headers=headers, timeout=60, verify=False)
            resp.raise_for_status()
            break
        except requests.exceptions.Timeout:
            if attempt == 2:
                return []
    raw = resp.json()
    if not raw:
        return []
    results = []
    for item in raw:
        cur = item.get("cur_unit", "").replace("(100)", "")
        if cur in TARGET_CURRENCIES:
            results.append({
                "currency": cur,
                "kftc_deal_bas_r": item.get("kftc_deal_bas_r", "").replace(",", ""),
            })
    return results


def fetch_lme_range(start: str, end: str) -> dict:
    """비철금속협회 LME 시세. {YYYYMMDD: {Cu, Ni}} 반환"""
    result = {}
    start_dt = datetime.strptime(start, "%Y%m%d")
    end_dt = datetime.strptime(end, "%Y%m%d")
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

    page = 1
    while True:
        url = f"https://www.nonferrous.or.kr/stats/?act=sub3&page={page}"
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            resp.raise_for_status()
        except Exception:
            break

        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table")
        if not table:
            break

        CU_IDX, NI_IDX = 1, 5
        oldest_on_page = None
        for row in table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) < 6:
                continue
            date_text = cells[0].get_text(strip=True)
            try:
                parts = [p.strip() for p in date_text.split(".") if p.strip()]
                date_str = "".join(parts)
                if len(date_str) != 8:
                    continue
                date_dt = datetime.strptime(date_str, "%Y%m%d")
            except Exception:
                continue
            oldest_on_page = date_dt
            if date_dt > end_dt or date_dt < start_dt:
                continue
            cu_val = cells[CU_IDX].get_text(strip=True) if CU_IDX < len(cells) else ""
            ni_val = cells[NI_IDX].get_text(strip=True) if NI_IDX < len(cells) else ""
            result[date_dt.strftime("%Y%m%d")] = {"Cu": cu_val, "Ni": ni_val}

        if oldest_on_page and oldest_on_page < start_dt:
            break
        if oldest_on_page is None:
            break
        page += 1
        time.sleep(0.2)

    return result


def fetch_opinet_range(start: str, end: str) -> dict:
    """OPINET 원유 가격. 월별 분할 + 월마다 세션 갱신 + 최대 14일 재시도.
    {YYYYMMDD: {Dubai, Brent, WTI}} 반환 ($/Bbl)"""
    result = {}
    start_dt = datetime.strptime(start, "%Y%m%d")
    end_dt   = datetime.strptime(end,   "%Y%m%d")

    url = "https://www.opinet.co.kr/glopcoilSelect.do"
    req_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer":    "https://www.opinet.co.kr/glopcoilSelect.do",
        "Accept":     "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    }

    def parse_opinet_date(text):
        """'26년04월01일' 또는 '2026년04월01일' → '20260401'"""
        m = re.search(r'(\d{2,4})년\s*(\d{1,2})월\s*(\d{1,2})일', text)
        if not m:
            return None
        yr = m.group(1)
        if len(yr) == 2:
            yr = "20" + yr
        return yr + m.group(2).zfill(2) + m.group(3).zfill(2)

    def extract_rows(resp_content):
        """tbody2 → div_dar 순으로 탐색, 인덱스 열 유무도 자동 감지"""
        soup = BeautifulSoup(resp_content, "html.parser")
        tbody = soup.find("tbody", id="tbody2")
        if not tbody:
            div = soup.find("div", id="div_dar")
            if div:
                tbody = div.find("tbody")
        if not tbody:
            return []
        rows_out = []
        for row in tbody.find_all("tr"):
            cells = row.find_all("td")
            # cells[0]=날짜, [1]=Dubai, [2]=Brent, [3]=WTI
            if len(cells) >= 4:
                date_str = parse_opinet_date(cells[0].get_text(strip=True))
                if date_str:
                    rows_out.append((date_str,
                                     cells[1].get_text(strip=True),
                                     cells[2].get_text(strip=True),
                                     cells[3].get_text(strip=True)))
                    continue
            # 인덱스 열이 앞에 있는 경우: cells[1]=날짜
            if len(cells) >= 5:
                date_str = parse_opinet_date(cells[1].get_text(strip=True))
                if date_str:
                    rows_out.append((date_str,
                                     cells[2].get_text(strip=True),
                                     cells[3].get_text(strip=True),
                                     cells[4].get_text(strip=True)))
        return rows_out

    current = start_dt.replace(day=1)
    while current <= end_dt:
        if current.month == 12:
            month_end = current.replace(year=current.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = current.replace(month=current.month + 1, day=1) - timedelta(days=1)

        chunk_end = min(month_end, end_dt)
        end_date  = chunk_end.strftime("%Y%m%d")

        session = requests.Session()
        session.headers.update(req_headers)
        try:
            session.get(url, timeout=15)
        except Exception:
            current = chunk_end + timedelta(days=1)
            continue

        found_rows = []
        for day_offset in range(14):
            try_dt = current + timedelta(days=day_offset)
            if try_dt.month != current.month:
                break
            std_date = try_dt.strftime("%Y%m%d")
            post_data = {
                "TERM": "D",
                "STA_Y": try_dt.strftime("%Y"), "STA_M": try_dt.strftime("%m"), "STA_D": try_dt.strftime("%d"),
                "END_Y": chunk_end.strftime("%Y"), "END_M": chunk_end.strftime("%m"), "END_D": chunk_end.strftime("%d"),
                "OILSRTCD1": "001", "OILSRTCD2": "002", "OILSRTCD3": "003",
                "STDDATE": std_date, "ENDDATE": end_date, "SEL_DIV": "div_dar",
            }
            try:
                resp = session.post(url, data=post_data, timeout=30)
                resp.raise_for_status()
            except Exception:
                time.sleep(0.15)
                continue
            rows = extract_rows(resp.content)
            if rows:
                found_rows = rows
                break
            time.sleep(0.15)

        for date_str, dubai, brent, wti in found_rows:
            try:
                row_dt = datetime.strptime(date_str, "%Y%m%d")
            except Exception:
                continue
            if row_dt < start_dt or row_dt > end_dt:
                continue
            if dubai == "0" and brent == "0" and wti == "0":
                continue
            result[date_str] = {"Dubai": dubai, "Brent": brent, "WTI": wti}

        current = chunk_end + timedelta(days=1)
        time.sleep(0.2)

    return result


def fetch_range_combined(start: str, end: str) -> list:
    fx_by_date = {}
    current = datetime.strptime(start, "%Y%m%d")
    end_dt = datetime.strptime(end, "%Y%m%d")
    while current <= end_dt:
        date_str = current.strftime("%Y%m%d")
        day = fetch_one_day_fx(date_str)
        if day:
            fx_by_date[date_str] = {item["currency"]: item["kftc_deal_bas_r"] for item in day}
        current += timedelta(days=1)
        time.sleep(0.1)

    lme_by_date   = fetch_lme_range(start, end)
    oil_by_date   = fetch_opinet_range(start, end)

    all_dates = sorted(set(list(fx_by_date.keys()) + list(lme_by_date.keys()) + list(oil_by_date.keys())))
    rows = []
    for date in all_dates:
        fx  = fx_by_date.get(date, {})
        lme = lme_by_date.get(date, {})
        oil = oil_by_date.get(date, {})
        rows.append({
            "date":  date,
            "USD":   fx.get("USD", ""),
            "EUR":   fx.get("EUR", ""),
            "Ni":    lme.get("Ni", ""),
            "Cu":    lme.get("Cu", ""),
            "Dubai": oil.get("Dubai", ""),
            "Brent": oil.get("Brent", ""),
            "WTI":   oil.get("WTI", ""),
        })
    return rows


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/range")
def api_range():
    try:
        start = request.args.get("start")
        end = request.args.get("end")
        if not start or not end:
            return jsonify({"error": "start, end 파라미터 필요"}), 400
        rows = fetch_range_combined(start, end)
        return jsonify({"rows": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download")
def api_download():
    try:
        start = request.args.get("start")
        end = request.args.get("end")
        rows = fetch_range_combined(start, end)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "환율_금속_유가"

        fx_fill   = PatternFill("solid", fgColor="1A3A5C")
        lme_fill  = PatternFill("solid", fgColor="3B1A5C")
        oil_fill  = PatternFill("solid", fgColor="1A4A2E")
        sfx_fill  = PatternFill("solid", fgColor="2D5F8A")
        slme_fill = PatternFill("solid", fgColor="6B2D8A")
        soil_fill = PatternFill("solid", fgColor="2D7A50")
        date_fill = PatternFill("solid", fgColor="2D4A6B")
        wfont  = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells("A1:A2"); ws["A1"] = "날짜"
        ws["A1"].fill = date_fill; ws["A1"].font = wfont; ws["A1"].alignment = center

        ws.merge_cells("B1:C1"); ws["B1"] = "환율 (KRW)"
        ws["B1"].fill = fx_fill; ws["B1"].font = wfont; ws["B1"].alignment = center

        ws.merge_cells("D1:E1"); ws["D1"] = "LME 현물 (US$/톤)"
        ws["D1"].fill = lme_fill; ws["D1"].font = wfont; ws["D1"].alignment = center

        ws.merge_cells("F1:H1"); ws["F1"] = "국제유가 ($/Bbl)"
        ws["F1"].fill = oil_fill; ws["F1"].font = wfont; ws["F1"].alignment = center

        sub_headers = [
            ("B", "USD/KRW", sfx_fill), ("C", "EUR/KRW", sfx_fill),
            ("D", "니켈 (Ni)", slme_fill), ("E", "구리 (Cu)", slme_fill),
            ("F", "두바이유", soil_fill), ("G", "브렌트", soil_fill), ("H", "WTI", soil_fill),
        ]
        for col, label, fill in sub_headers:
            cell = ws[col + "2"]
            cell.value = label; cell.fill = fill
            cell.font = wfont; cell.alignment = center; cell.border = border
        ws["A2"].border = border

        def to_float(val):
            try:
                return float(str(val).replace(",", ""))
            except Exception:
                return None

        for r in rows:
            date_fmt = r["date"][:4] + "-" + r["date"][4:6] + "-" + r["date"][6:8]
            ws.append([date_fmt,
                       to_float(r["USD"]), to_float(r["EUR"]),
                       to_float(r["Ni"]),  to_float(r["Cu"]),
                       to_float(r["Dubai"]), to_float(r["Brent"]), to_float(r["WTI"])])
            row_idx = ws.max_row
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx == 1:
                    cell.alignment = center
                else:
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.00"

        for col, width in [("A",14),("B",13),("C",13),("D",14),("E",14),("F",12),("G",10),("H",10)]:
            ws.column_dimensions[col].width = width
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        filename = f"환율_금속_유가_{start}_{end}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    print(f"서버 시작: http://localhost:{port}")
    app.run(host="0.0.0.0", debug=False, port=port)
